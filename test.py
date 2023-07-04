import unittest
from unittest.mock import Mock, patch
from openpyxl import Workbook
from datetime import datetime
from io import BytesIO

# Import the functions to be tested
from script import get_package_manager, get_dependency_management, get_semantic_release, get_gha, get_workflow_info, update_excel, process_repo


class TestGitHubRepoProcessing(unittest.TestCase):
    def setUp(self):
        # Create a mock repository object for testing
        self.repo = Mock()
        self.repo.name = "sample-repo"
        self.repo.get_contents.return_value = []
        self.repo.get_pulls.return_value.totalCount = 0

        # Create an in-memory Excel workbook for testing
        self.workbook = Workbook()
        self.sheet = self.workbook.active

        # Patch the openpyxl functions used in the code
        patcher = patch("<module_name>.openpyxl.load_workbook", return_value=self.workbook)
        self.mock_load_workbook = patcher.start()
        self.addCleanup(patcher.stop)

    def test_get_package_manager_with_package_lock(self):
        # Mock the repository contents to include 'package-lock.json'
        self.repo.get_contents.return_value = [Mock(name="package-lock.json")]

        # Call the get_package_manager function
        package_manager = get_package_manager(self.repo)

        # Assert the expected package_manager value
        self.assertEqual(package_manager, "NPM")

    def test_get_package_manager_with_yarn_lock(self):
        # Mock the repository contents to include 'yarn.lock'
        self.repo.get_contents.return_value = [Mock(name="yarn.lock")]

        # Call the get_package_manager function
        package_manager = get_package_manager(self.repo)

        # Assert the expected package_manager value
        self.assertEqual(package_manager, "Yarn")

    def test_get_package_manager_without_lock_files(self):
        # Mock the repository contents to exclude lock files
        self.repo.get_contents.return_value = []

        # Call the get_package_manager function
        package_manager = get_package_manager(self.repo)

        # Assert the expected package_manager value
        self.assertEqual(package_manager, "No")

    def test_get_dependency_management_with_dependabot(self):
        # Mock the repository to have Dependabot pull requests
        self.repo.get_pulls.return_value.totalCount = 2
        self.repo.get_pulls.return_value[0].creator.login = "dependabot"

        # Call the get_dependency_management function
        dependency_management = get_dependency_management(self.repo)

        # Assert the expected dependency_management value
        self.assertEqual(dependency_management, "Dependabot")

    def test_get_dependency_management_with_renovate(self):
        # Mock the repository to have Renovate pull requests
        self.repo.get_pulls.return_value.totalCount = 1
        self.repo.get_pulls.return_value[0].creator.login = "renovate"

        # Call the get_dependency_management function
        dependency_management = get_dependency_management(self.repo)

        # Assert the expected dependency_management value
        self.assertEqual(dependency_management, "Renovate")

    def test_get_dependency_management_without_management(self):
        # Mock the repository to have no dependency management pull requests
        self.repo.get_pulls.return_value.totalCount = 0

        # Call the get_dependency_management function
        dependency_management = get_dependency_management(self.repo)

        # Assert the expected dependency_management value
        self.assertEqual(dependency_management, "No")

    def test_get_semantic_release_with_semantic_release_deps(self):
        # Mock the repository to have 'semantic-release' in devDependencies
        self.repo.get_contents.return_value = [Mock(name="package.json", decoded_content=b'{"devDependencies": {"semantic-release": "^1.0.0"}}')]

        # Call the get_semantic_release function
        semantic_release = get_semantic_release(self.repo)

        # Assert the expected semantic_release value
        self.assertEqual(semantic_release, "Yes")

    def test_get_semantic_release_without_semantic_release_deps(self):
        # Mock the repository to not have 'semantic-release' in devDependencies
        self.repo.get_contents.return_value = [Mock(name="package.json", decoded_content=b'{"devDependencies": {}}')]

        # Call the get_semantic_release function
        semantic_release = get_semantic_release(self.repo)

        # Assert the expected semantic_release value
        self.assertEqual(semantic_release, "No")

    def test_get_gha_with_recent_workflow_runs(self):
        # Mock the repository to have recent workflow runs
        workflow = Mock()
        workflow.last_modified = datetime.now()

        self.repo.get_contents.return_value = [workflow]

        # Call the get_gha function
        gha = get_gha(self.repo.get_contents(".github/workflows"))

        # Assert the expected gha value
        self.assertEqual(gha, "Yes")

    def test_get_gha_with_no_recent_workflow_runs(self):
        # Mock the repository to not have recent workflow runs
        workflow = Mock()
        workflow.last_modified = datetime(2022, 1, 1)

        self.repo.get_contents.return_value = [workflow]

        # Call the get_gha function
        gha = get_gha(self.repo.get_contents(".github/workflows"))

        # Assert the expected gha value
        self.assertEqual(gha, "No")

    def test_get_workflow_info(self):
        # Mock the repository workflows and their content
        workflows = [
            Mock(name="workflow1.yml", path=".github/workflows/workflow1.yml", decoded_content=b"content1"),
            Mock(name="workflow2.yml", path=".github/workflows/workflow2.yml", decoded_content=b"content2"),
        ]

        # Patch the workflow YAML parsing to return predefined values
        with patch("<module_name>.yaml.safe_load") as mock_yaml_load:
            mock_yaml_load.side_effect = [
                {"env": {"INTEGRATION_SUITE": "Yes"}, "concurrency": {"group": "Yes"}, "steps": {"mend": "Yes"}},
                {"env": {}, "concurrency": {}, "steps": {}},
            ]

            # Call the get_workflow_info function
            workflow_info = get_workflow_info(workflows)

            # Assert the expected workflow_info values
            self.assertEqual(
                workflow_info,
                {
                    "workflow1.yml": {"Integration Suite": "Yes", "Concurrency Rule": "Yes", "Mend": "Yes"},
                    "workflow2.yml": {"Integration Suite": None, "Concurrency Rule": None, "Mend": None},
                },
            )

    def test_update_excel_existing_repo(self):
        # Create a sample Excel sheet with an existing repository row
        self.sheet.append(["sample-repo", "", "", "No", "No", "No", "No", None, None, None])

        # Call the update_excel function
        update_excel("sample-repo", "NPM", "Renovate", "Yes", "Yes", "Yes", "Yes", "Yes")

        # Retrieve the updated row values
        row_values = [cell.value for cell in self.sheet[2]]

        # Assert the expected updated row values
        self.assertEqual(row_values, ["sample-repo", "", "", "NPM", "Renovate", "Yes", "Yes", "Yes", "Yes", "Yes"])

    def test_update_excel_new_repo(self):
        # Call the update_excel function for a new repository
        update_excel("new-repo", "Yarn", "Dependabot", "No", "No", "No", "No", "No")

        # Retrieve the updated row values
        row_values = [cell.value for cell in self.sheet[2]]

        # Assert the expected updated row values
        self.assertEqual(row_values, ["new-repo", "", "", "Yarn", "Dependabot", "No", "No", "No", "No", "No"])

    def test_process_repo_success(self):
        # Patch the necessary functions to return predefined values
        with patch("<module_name>.get_package_manager", return_value="NPM") as mock_get_package_manager, \
                patch("<module_name>.get_dependency_management", return_value="Renovate") as mock_get_dependency_management, \
                patch("<module_name>.get_semantic_release", return_value="Yes") as mock_get_semantic_release, \
                patch("<module_name>.get_gha", return_value="Yes") as mock_get_gha, \
                patch("<module_name>.get_workflow_info", return_value={"workflow.yml": {}}) as mock_get_workflow_info, \
                patch("<module_name>.update_excel") as mock_update_excel:

            # Call the process_repo function
            process_repo(self.repo)

            # Assert that the necessary functions were called with the expected arguments
            mock_get_package_manager.assert_called_once_with(self.repo)
            mock_get_dependency_management.assert_called_once_with(self.repo)
            mock_get_semantic_release.assert_called_once_with(self.repo)
            mock_get_gha.assert_called_once_with([])
            mock_get_workflow_info.assert_called_once_with([])
            mock_update_excel.assert_called_once_with(
                "sample-repo", "NPM", "Renovate", "Yes", "Yes", {}, None, None
            )

    def test_process_repo_error(self):
        # Patch the necessary functions to raise an exception
        with patch("<module_name>.get_package_manager", side_effect=Exception("Error processing package manager")), \
                patch("<module_name>.get_dependency_management", return_value="Renovate"), \
                patch("<module_name>.get_semantic_release", return_value="Yes"), \
                patch("<module_name>.get_gha", return_value="Yes"), \
                patch("<module_name>.get_workflow_info", return_value={"workflow.yml": {}}), \
                patch("<module_name>.update_excel"):

            # Call the process_repo function and catch the exception
            with self.assertRaises(Exception) as cm:
                process_repo(self.repo)

            # Assert that the exception message contains the repository name
            self.assertIn("sample-repo", str(cm.exception))

    def test_full_processing_workflow(self):
        # Create a sample repository with predefined values
        repo = Mock()
        repo.name = "sample-repo"
        repo.get_contents.return_value = [Mock(name="package-lock.json")]
        repo.get_pulls.return_value.totalCount = 1
        repo.get_pulls.return_value[0].creator.login = "dependabot"

        # Patch the necessary functions to return predefined values
        with patch("<module_name>.Github", return_value=Mock(get_repo=lambda name: repo)), \
                patch("<module_name>.get_workflow_info", return_value={"workflow.yml": {}}) as mock_get_workflow_info, \
                patch("<module_name>.process_repo") as mock_process_repo:

            # Call the main function to process the repository
            main_function()

            # Assert that the necessary functions were called with the expected arguments
            mock_get_workflow_info.assert_called_once_with([])
            mock_process_repo.assert_called_once_with(repo)

    def test_full_processing_workflow_with_error(self):
        # Create a sample repository with an exception-raising get_package_manager function
        repo = Mock()
        repo.name = "sample-repo"
        repo.get_package_manager.side_effect = Exception("Error processing package manager")

        # Patch the necessary functions to return predefined values
        with patch("<module_name>.Github", return_value=Mock(get_repo=lambda name: repo)), \
                patch("<module_name>.get_workflow_info", return_value={"workflow.yml": {}}), \
                patch("<module_name>.process_repo"):

            # Call the main function to process the repository and catch the exception
            with self.assertRaises(Exception) as cm:
                main_function()

            # Assert that the exception message contains the repository name
            self.assertIn("sample-repo", str(cm.exception))


if __name__ == "__main__":
    unittest.main()
